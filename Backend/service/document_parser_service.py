import json
import chardet
from typing import Tuple, Dict, Any
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

def detect_file_type(filename: str, content: bytes) -> str:
    """
    Detect file type based on filename extension and content analysis
    """
    filename_lower = filename.lower()
    
    if filename_lower.endswith('.json'):
        return 'json'
    elif filename_lower.endswith('.pdf'):
        return 'pdf'
    elif filename_lower.endswith(('.doc', '.docx')):
        return 'word'
    elif filename_lower.endswith(('.xls', '.xlsx')):
        return 'excel'
    elif filename_lower.endswith('.txt'):
        return 'text'
    else:
        # Try to detect by content
        try:
            # Try to decode as text first
            encoding = chardet.detect(content)['encoding']
            if encoding:
                decoded = content.decode(encoding)
                json.loads(decoded)
                return 'json'
        except:
            pass
        return 'unknown'

def parse_json_content(content: bytes) -> Dict[Any, Any]:
    """Parse JSON content with encoding detection"""
    try:
        # First try UTF-8
        decoded = content.decode('utf-8')
        return json.loads(decoded)
    except UnicodeDecodeError:
        # If UTF-8 fails, detect encoding
        encoding = chardet.detect(content)['encoding']
        if encoding:
            decoded = content.decode(encoding)
            return json.loads(decoded)
        else:
            raise ValueError("Could not detect file encoding")
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON format: {e}")

def parse_text_content(content: bytes) -> str:
    """Parse plain text content with encoding detection"""
    try:
        # First try UTF-8
        return content.decode('utf-8')
    except UnicodeDecodeError:
        # If UTF-8 fails, detect encoding
        encoding = chardet.detect(content)['encoding']
        if encoding:
            return content.decode(encoding)
        else:
            raise ValueError("Could not detect file encoding")

def parse_pdf_content(content: bytes) -> str:
    """Parse PDF content and extract text"""
    try:
        import PyPDF2
        
        pdf_file = BytesIO(content)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        
        text_content = []
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text_content.append(page.extract_text())
        
        return '\n'.join(text_content)
    except ImportError:
        raise ValueError("PyPDF2 library not installed. Cannot parse PDF files.")
    except Exception as e:
        raise ValueError(f"Error parsing PDF: {e}")

def parse_word_content(content: bytes) -> str:
    """Parse Word document content and extract text"""
    try:
        from docx import Document
        
        doc_file = BytesIO(content)
        doc = Document(doc_file)
        
        text_content = []
        for paragraph in doc.paragraphs:
            text_content.append(paragraph.text)
        
        return '\n'.join(text_content)
    except ImportError:
        raise ValueError("python-docx library not installed. Cannot parse Word documents.")
    except Exception as e:
        raise ValueError(f"Error parsing Word document: {e}")

def parse_excel_content(content: bytes) -> str:
    """Parse Excel content and extract text from all sheets"""
    try:
        import openpyxl
        
        excel_file = BytesIO(content)
        workbook = openpyxl.load_workbook(excel_file)
        
        text_content = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_content.append(f"Sheet: {sheet_name}")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        row_text.append(str(cell))
                if row_text:
                    text_content.append('\t'.join(row_text))
        
        return '\n'.join(text_content)
    except ImportError:
        raise ValueError("openpyxl library not installed. Cannot parse Excel files.")
    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {e}")

def parse_document_content(filename: str, content: bytes) -> Tuple[str, Any]:
    """
    Parse document content based on file type
    Returns: (file_type, parsed_content)
    """
    file_type = detect_file_type(filename, content)
    logger.info(f"Detected file type: {file_type} for file: {filename}")
    
    try:
        if file_type == 'json':
            parsed_content = parse_json_content(content)
            return file_type, parsed_content
        
        elif file_type == 'pdf':
            text_content = parse_pdf_content(content)
            # For transcripts, wrap text content in expected JSON structure
            parsed_content = {
                "transcript": text_content,
                "metadata": {
                    "source": filename,
                    "type": "pdf_extracted"
                }
            }
            return file_type, parsed_content
        
        elif file_type == 'word':
            text_content = parse_word_content(content)
            # For transcripts, wrap text content in expected JSON structure
            parsed_content = {
                "transcript": text_content,
                "metadata": {
                    "source": filename,
                    "type": "word_extracted"
                }
            }
            return file_type, parsed_content
        
        elif file_type == 'excel':
            text_content = parse_excel_content(content)
            # For transcripts, wrap text content in expected JSON structure
            parsed_content = {
                "transcript": text_content,
                "metadata": {
                    "source": filename,
                    "type": "excel_extracted"
                }
            }
            return file_type, parsed_content
        
        elif file_type == 'text':
            text_content = parse_text_content(content)
            # For transcripts, wrap text content in expected JSON structure
            parsed_content = {
                "transcript": text_content,
                "metadata": {
                    "source": filename,
                    "type": "text_file"
                }
            }
            return file_type, parsed_content
        
        else:
            # Try to parse as text as fallback
            try:
                text_content = parse_text_content(content)
                parsed_content = {
                    "transcript": text_content,
                    "metadata": {
                        "source": filename,
                        "type": "text_fallback"
                    }
                }
                return 'text', parsed_content
            except:
                raise ValueError(f"Unsupported file type: {file_type}")
    
    except Exception as e:
        logger.error(f"Error parsing document {filename}: {e}")
        raise ValueError(f"Error parsing document: {e}")
